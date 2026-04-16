from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from flask_pymongo import PyMongo
from flask_jwt_extended import JWTManager, create_access_token, jwt_required, get_jwt_identity
from werkzeug.security import generate_password_hash, check_password_hash
from bson.objectid import ObjectId
import pandas as pd
from io import BytesIO
import os
from cryptography.fernet import Fernet
import re
import base64
from datetime import datetime, timedelta
import traceback
import requests
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address

# ============== Environment & Security Configuration ==============
from dotenv import load_dotenv
load_dotenv()


# Kritik environment variable kontrolü
def get_required_env(key):
    value = os.environ.get(key)
    if not value:
        raise ValueError(f"{key} environment variable must be set")
    return value

# ============== Flask Yapılandırması ==============
app = Flask(__name__)
app.config['SECRET_KEY'] = get_required_env('SECRET_KEY')
app.config['MONGO_URI'] = get_required_env('MONGO_URI')
app.config['JWT_SECRET_KEY'] = get_required_env('JWT_SECRET_KEY')
app.config['JWT_ACCESS_TOKEN_EXPIRES'] = timedelta(days=7)
app.config['MAX_CONTENT_LENGTH'] = int(os.environ.get('MAX_CONTENT_LENGTH', 10 * 1024 * 1024))

# Encryption key kontrolü ve yapılandırması
ENCRYPTION_KEY = get_required_env('ENCRYPTION_KEY')
try:
    ENCRYPTION_KEY = ENCRYPTION_KEY.encode()
    cipher_suite = Fernet(ENCRYPTION_KEY)
except Exception as e:
    raise ValueError(f"Invalid ENCRYPTION_KEY format: {str(e)}")

# CORS Yapılandırması - Wildcard kaldırıldı
ALLOWED_ORIGINS = os.environ.get('ALLOWED_ORIGINS', 'http://localhost:4200,https://csv-convert-front.onrender.com')
ALLOWED_ORIGINS = [origin.strip() for origin in ALLOWED_ORIGINS.split(',')]

CORS(app, resources={
    r"/api/*": {
        "origins": ALLOWED_ORIGINS,
        "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
        "allow_headers": ["Content-Type", "Authorization", "X-Clockify-Api-Key"],
        "supports_credentials": True,
        "expose_headers": ["Content-Type", "Authorization"]
    }
})

# Rate Limiting Yapılandırması - Memory-based (Redis yok)
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=[os.environ.get('DEFAULT_RATE_LIMIT', '200 per day,50 per hour,10 per minute')],
    storage_uri='memory://'  # Memory-based storage
)

# MongoDB & JWT
mongo = PyMongo(app)
jwt = JWTManager(app)

# File upload güvenliği
ALLOWED_EXTENSIONS = {'csv'}
ALLOWED_MIME_TYPES = {'text/csv', 'text/plain', 'application/vnd.ms-excel'}

# Input validation regex
VALID_NAME_RE = re.compile(r'^[\w\s\-\.@]+$')
VALID_EMAIL_RE = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')

# Password validation regex (min 8 karakter, büyük, küçük, rakam, özel karakter)
PASSWORD_REGEX = re.compile(r'^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&])[A-Za-z\d@$!%*?&]{8,}$')

# ============== Security Headers Middleware ==============

@app.after_request
def set_security_headers(response):
    """Güvenlik header'larını tüm response'lara ekle"""
    # CORS headers'ları manuel olarak da ekle
    origin = request.headers.get('Origin')
    if origin in ALLOWED_ORIGINS:
        response.headers['Access-Control-Allow-Origin'] = origin
        response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization, X-Clockify-Api-Key'
        response.headers['Access-Control-Allow-Credentials'] = 'true'
    
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Permissions-Policy'] = 'geolocation=(), microphone=(), camera=()'
    
    # HSTS header (HTTPS zorunlu)
    if os.environ.get('FORCE_HTTPS', 'True') == 'True':
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    
    return response

@app.route('/api/<path:path>', methods=['OPTIONS'])
def handle_options(path):
    """Handle preflight OPTIONS requests"""
    response = jsonify({'status': 'ok'})
    origin = request.headers.get('Origin')
    if origin in ALLOWED_ORIGINS:
        response.headers['Access-Control-Allow-Origin'] = origin
        response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization, X-Clockify-Api-Key'
        response.headers['Access-Control-Allow-Credentials'] = 'true'
    return response, 200

# ============== Yardımcı Fonksiyonlar ==============

def validate_password(password):
    """Şifre güvenlik kontrolü"""
    if not password or len(password) < 8:
        return False, "Password must be at least 8 characters long"
    
    if not PASSWORD_REGEX.match(password):
        return False, "Password must contain uppercase, lowercase, number and special character"
    
    return True, "Valid"

def validate_email(email):
    """Email format kontrolü"""
    if not email or not VALID_EMAIL_RE.match(email):
        return False, "Invalid email format"
    return True, "Valid"

def sanitize_input(text, max_length=255):
    """Input sanitization - XSS ve injection koruması"""
    if not text:
        return ""
    
    text = str(text)
    
    if len(text) > max_length:
        text = text[:max_length]
    
    text = text.replace('<', '&lt;').replace('>', '&gt;')
    text = text.replace('"', '&quot;').replace("'", '&#x27;')
    
    return text.strip()

def sanitize_mongodb_query(query_dict):
    """MongoDB injection koruması"""
    if not isinstance(query_dict, dict):
        return {}
    
    safe_query = {}
    for key, value in query_dict.items():
        if not key.startswith('$'):
            if isinstance(value, dict):
                safe_query[key] = sanitize_mongodb_query(value)
            else:
                safe_query[key] = value
    
    return safe_query

def validate_file(file):
    """Dosya güvenlik kontrolü"""
    if not file or file.filename == '':
        return False, "No file selected"
    
    if '.' not in file.filename:
        return False, "Invalid file format"
    
    extension = file.filename.rsplit('.', 1)[1].lower()
    if extension not in ALLOWED_EXTENSIONS:
        return False, f"File type not allowed. Allowed: {', '.join(ALLOWED_EXTENSIONS)}"
    
    file.seek(0, 2)
    size = file.tell()
    file.seek(0)
    
    if size > app.config['MAX_CONTENT_LENGTH']:
        return False, f"File too large. Max size: {app.config['MAX_CONTENT_LENGTH'] / (1024*1024)}MB"
    
    return True, "Valid"

def encrypt_api_key(api_key):
    """API key'i şifrele"""
    if not api_key:
        return None
    return cipher_suite.encrypt(api_key.encode()).decode()

def decrypt_api_key(encrypted_key):
    """API key'i çöz"""
    if not encrypted_key:
        return None
    try:
        return cipher_suite.decrypt(encrypted_key.encode()).decode()
    except:
        return None

def safe_error_response(error, status_code=500):
    """Güvenli hata mesajı döndür - internal detay verme"""
    app.logger.error(f"Error: {str(error)}\n{traceback.format_exc()}")
    
    error_messages = {
        400: "Bad request. Please check your input.",
        401: "Authentication required.",
        403: "Access denied.",
        404: "Resource not found.",
        500: "An internal error occurred. Please try again later."
    }
    
    return jsonify({
        'error': error_messages.get(status_code, "An error occurred"),
        'status': status_code
    }), status_code

def parse_date_range(date_str):
    """
    JS .toISOString() UTC'ye çevirince Türkiye (UTC+3) için
    01-05-2025 00:00 lokal → 30-04-2025 21:00 UTC oluyor.
    Sadece DATE kısmını (YYYY-MM-DD) alarak timezone kaymasını önler.
    """
    if not date_str:
        return None
    try:
        date_only = str(date_str)[:10]  # "2025-05-01T21:00:00Z" → "2025-05-01"
        return pd.Timestamp(date_only)
    except Exception:
        return None
 
def _safe_str(value, fallback):
    """
    None, boş string ve yalnızca boşluktan oluşan değerleri
    fallback ile değiştirir. str(None) → "None" sorununu önler.
    """
    if value is None:
        return fallback
    v = str(value).strip()
    return v if v else fallback
 
 
def _get_range(date_range_start, date_range_end, df):
    """Tarih aralığını timezone-safe şekilde hesapla."""
    if date_range_start and date_range_end:
        rs = parse_date_range(date_range_start)
        re_ = parse_date_range(date_range_end)
        if rs is not None and re_ is not None:
            return rs, re_
    if not df['ParsedDate'].dropna().empty:
        return df['ParsedDate'].min(), df['ParsedDate'].max()
    return pd.Timestamp("2025-01-01"), pd.Timestamp("2025-01-31")

# ============== Excel Generation Functions ==============

def parse_duration_to_seconds(d_str):
    try:
        h, m, s = map(int, d_str.split(":"))
        return h * 3600 + m * 60 + s
    except:
        return 0

def round_to_nearest_minute(seconds):
    minutes = seconds / 60
    rounded_minutes = round(minutes)
    return rounded_minutes * 60

def sanitize_excel_cell(value):
    if pd.isna(value) or value is None:
        return ""
    value = str(value)
    if value and value[0] in ('=', '+', '-', '@'):
        return "'" + value
    return value

def generate_excel_report(df, format_choice, report_period, projects, customers, logo_data=None, company_info=None, date_range_start=None, date_range_end=None):
    """Excel raporu oluşturur - Summary ve Detailed Report ile"""
    output = BytesIO()
 
    df["raw_seconds"] = df["Duration (h)"].apply(parse_duration_to_seconds)
    df["rounded_seconds"] = df["raw_seconds"].apply(round_to_nearest_minute)
 
    if format_choice == "hours":
        df["formatted_duration"] = df["rounded_seconds"] / 86400
    else:
        df["formatted_duration"] = (df["rounded_seconds"] / 3600).round(2)
 
    df['Start Date'] = df['Start Date'].astype(str)
    df['ParsedDate'] = pd.to_datetime(df['Start Date'], format='%d/%m/%Y', errors='coerce')
    df["Day"] = df["ParsedDate"].apply(lambda d: d.strftime("%d (%A)") if pd.notnull(d) else "Unknown")
    df["DayFull"] = df["ParsedDate"].apply(lambda d: d.strftime("%d %B %Y (%A)") if pd.notnull(d) else "Unknown")
 
    if "Start Time" not in df.columns:
        df["Start Time"] = ""
    if "End Time" not in df.columns:
        df["End Time"] = ""
 
    # ── DÜZELTİLMİŞ: timezone-safe tarih aralığı hesabı ──
    range_start, range_end = _get_range(date_range_start, date_range_end, df)
 
    all_days = pd.date_range(start=range_start, end=range_end, freq='D')
    all_days_str = [d.strftime("%d (%A)") for d in all_days]
    all_days_full = [d.strftime("%d %B %Y (%A)") for d in all_days]
 
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        workbook = writer.book
 
        header_format = workbook.add_format({
            'bold': True, 'border': 1, 'bg_color': '#4472C4',
            'font_color': 'white', 'align': 'center', 'valign': 'vcenter'
        })
        info_label_format = workbook.add_format({
            'bold': True, 'border': 1, 'bg_color': '#4472C4',
            'font_color': 'white', 'align': 'left', 'valign': 'vcenter'
        })
        info_value_format = workbook.add_format({
            'border': 1, 'align': 'left', 'valign': 'vcenter', 'text_wrap': True
        })
        cell_format = workbook.add_format({'border': 1, 'align': 'left', 'valign': 'vcenter'})
        cell_wrap_format = workbook.add_format({'border': 1, 'align': 'left', 'valign': 'top', 'text_wrap': True})
        cell_center_format = workbook.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter'})
        number_format = workbook.add_format({'num_format': '0.00', 'border': 1, 'align': 'right', 'valign': 'vcenter'})
        time_format = workbook.add_format({'num_format': '[h]:mm', 'border': 1, 'align': 'right', 'valign': 'vcenter'})
        yellow_format = workbook.add_format({'bg_color': '#FFEB9C', 'border': 1, 'bold': True, 'align': 'center', 'valign': 'vcenter'})
        yellow_number_format = workbook.add_format({'bg_color': '#FFEB9C', 'border': 1, 'bold': True, 'num_format': '0.00', 'align': 'right', 'valign': 'vcenter'})
        yellow_time_format = workbook.add_format({'bg_color': '#FFEB9C', 'border': 1, 'bold': True, 'num_format': '[h]:mm', 'align': 'right', 'valign': 'vcenter'})
        green_format = workbook.add_format({'bg_color': '#C6EFCE', 'border': 1, 'bold': True, 'align': 'center', 'valign': 'vcenter', 'font_color': '#006100'})
        green_number_format = workbook.add_format({'bg_color': '#C6EFCE', 'border': 1, 'bold': True, 'num_format': '0.00', 'align': 'right', 'valign': 'vcenter', 'font_color': '#006100'})
        green_time_format = workbook.add_format({'bg_color': '#C6EFCE', 'border': 1, 'bold': True, 'num_format': '[h]:mm', 'align': 'right', 'valign': 'vcenter', 'font_color': '#006100'})
        red_format = workbook.add_format({'bg_color': '#FFC7CE', 'border': 1, 'bold': True, 'align': 'center', 'valign': 'vcenter', 'font_color': '#9C0006'})
        red_number_format = workbook.add_format({'bg_color': '#FFC7CE', 'border': 1, 'bold': True, 'num_format': '0.00', 'align': 'right', 'valign': 'vcenter', 'font_color': '#9C0006'})
        red_time_format = workbook.add_format({'bg_color': '#FFC7CE', 'border': 1, 'bold': True, 'num_format': '[h]:mm', 'align': 'right', 'valign': 'vcenter', 'font_color': '#9C0006'})
        user_header_format = workbook.add_format({'bold': True, 'border': 1, 'bg_color': '#4472C4', 'font_color': 'white', 'align': 'left', 'valign': 'vcenter', 'font_size': 12})
        empty_day_format = workbook.add_format({'border': 1, 'align': 'left', 'valign': 'vcenter', 'bg_color': '#F2F2F2', 'font_color': '#AAAAAA'})
        empty_cell_format = workbook.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter', 'bg_color': '#F2F2F2', 'font_color': '#AAAAAA'})
        empty_number_format = workbook.add_format({'num_format': '0.00', 'border': 1, 'align': 'right', 'valign': 'vcenter', 'bg_color': '#F2F2F2', 'font_color': '#AAAAAA'})
        empty_time_format = workbook.add_format({'num_format': '[h]:mm', 'border': 1, 'align': 'right', 'valign': 'vcenter', 'bg_color': '#F2F2F2', 'font_color': '#AAAAAA'})
        detail_header_format = workbook.add_format({'bold': True, 'border': 1, 'bg_color': '#667eea', 'font_color': 'white', 'align': 'center', 'valign': 'vcenter', 'font_size': 11})
        detail_date_header_format = workbook.add_format({'bold': True, 'border': 1, 'bg_color': '#4472C4', 'font_color': 'white', 'align': 'left', 'font_size': 12, 'valign': 'vcenter'})
        detail_date_empty_format = workbook.add_format({'bold': True, 'border': 1, 'bg_color': '#AAAAAA', 'font_color': '#FFFFFF', 'align': 'left', 'font_size': 12, 'valign': 'vcenter'})
        detail_cell_format = workbook.add_format({'border': 1, 'align': 'left', 'valign': 'vcenter', 'font_size': 10})
        detail_cell_wrap = workbook.add_format({'border': 1, 'align': 'left', 'valign': 'top', 'font_size': 10, 'text_wrap': True})
        detail_cell_center = workbook.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter', 'font_size': 10})
        detail_number_format = workbook.add_format({'num_format': '0.00', 'border': 1, 'align': 'right', 'valign': 'vcenter', 'font_size': 10})
        detail_time_format = workbook.add_format({'num_format': '[h]:mm', 'border': 1, 'align': 'right', 'valign': 'vcenter', 'font_size': 10})
        detail_empty_cell = workbook.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter', 'font_size': 10, 'bg_color': '#F2F2F2', 'font_color': '#AAAAAA'})
        detail_empty_wrap = workbook.add_format({'border': 1, 'align': 'left', 'valign': 'vcenter', 'font_size': 10, 'bg_color': '#F2F2F2', 'font_color': '#AAAAAA'})
 
        # ============== SAYFA 1: ÖZET RAPOR ==============
        summary_sheet = workbook.add_worksheet("Summary Report")
        summary_sheet.fit_to_pages(1, 0)
        summary_sheet.set_landscape()
        summary_sheet.set_paper(9)
 
        row = 0
 
        if company_info or logo_data:
            table_start_row = row
 
            if company_info:
                summary_sheet.write(row, 0, "Company:", info_label_format)
                summary_sheet.merge_range(row, 1, row, 2, sanitize_excel_cell(company_info.get('company_name', '')), info_value_format)
                summary_sheet.set_row(row, 18)
                row += 1
 
                if company_info.get('contact_person'):
                    summary_sheet.write(row, 0, "Contact:", info_label_format)
                    summary_sheet.merge_range(row, 1, row, 2, sanitize_excel_cell(company_info.get('contact_person', '')), info_value_format)
                    summary_sheet.set_row(row, 18)
                    row += 1
 
                if company_info.get('phone'):
                    summary_sheet.write(row, 0, "Phone:", info_label_format)
                    summary_sheet.merge_range(row, 1, row, 2, sanitize_excel_cell(company_info.get('phone', '')), info_value_format)
                    summary_sheet.set_row(row, 18)
                    row += 1
 
                if company_info.get('address'):
                    summary_sheet.write(row, 0, "Address:", info_label_format)
                    summary_sheet.merge_range(row, 1, row, 2, sanitize_excel_cell(company_info.get('address', '')), info_value_format)
                    summary_sheet.set_row(row, 18)
                    row += 1
 
            if logo_data is not None:
                try:
                    from PIL import Image as PILImage
                    pil_img = PILImage.open(BytesIO(logo_data['data']))
                    pil_img.thumbnail((180, 80), PILImage.LANCZOS)
                    opt_io = BytesIO()
                    pil_img.save(opt_io, format='PNG')
                    opt_io.seek(0)
                    logo_end_row = max(table_start_row, row - 1)
                    summary_sheet.merge_range(table_start_row, 3, logo_end_row, 4, "", info_value_format)
                    summary_sheet.insert_image(
                        table_start_row, 3, "logo.png",
                        {'image_data': opt_io, 'x_offset': 0, 'y_offset': 0, 'x_scale': 1.85, 'y_scale': 1.2, 'positioning': 1}
                    )
                except:
                    pass
 
            row += 1
 
        summary_sheet.write(row, 0, "Period:", info_label_format)
        summary_sheet.merge_range(row, 1, row, 4, sanitize_excel_cell(report_period), info_value_format)
        summary_sheet.set_row(row, 18)
        row += 1
 
        summary_sheet.write(row, 0, "Projects:", info_label_format)
        summary_sheet.merge_range(row, 1, row, 4, sanitize_excel_cell(projects), info_value_format)
        summary_sheet.set_row(row, 18)
        row += 1
 
        summary_sheet.write(row, 0, "Customers:", info_label_format)
        summary_sheet.merge_range(row, 1, row, 4, sanitize_excel_cell(customers), info_value_format)
        summary_sheet.set_row(row, 18)
        row += 2
 
        for user in sorted(df["User"].dropna().unique()):
            user_df = df[df["User"] == user].copy()
 
            summary_sheet.merge_range(row, 0, row, 4, sanitize_excel_cell(f"User: {user}"), user_header_format)
            summary_sheet.set_row(row, 20)
            row += 1
 
            summary_sheet.write(row, 0, "Day", header_format)
            summary_sheet.write(row, 1, "Description", header_format)
            summary_sheet.write(row, 2, "Billable Duration", header_format)
            summary_sheet.write(row, 3, "Free Duration", header_format)
            summary_sheet.write(row, 4, "Total Duration", header_format)
            summary_sheet.set_row(row, 18)
            row += 1
 
            for day_str in all_days_str:
                day_df = user_df[user_df["Day"] == day_str]
 
                if day_df.empty:
                    summary_sheet.write(row, 0, sanitize_excel_cell(day_str), empty_day_format)
                    summary_sheet.write(row, 1, "-", empty_cell_format)
                    if format_choice == "hours":
                        summary_sheet.write_number(row, 2, 0, empty_time_format)
                        summary_sheet.write_number(row, 3, 0, empty_time_format)
                        summary_sheet.write_number(row, 4, 0, empty_time_format)
                    else:
                        summary_sheet.write_number(row, 2, 0.0, empty_number_format)
                        summary_sheet.write_number(row, 3, 0.0, empty_number_format)
                        summary_sheet.write_number(row, 4, 0.0, empty_number_format)
                    summary_sheet.set_row(row, 18)
                else:
                    unique_descriptions = []
                    for desc in day_df["Description"].tolist():
                        desc_str = str(desc).strip()
                        if desc_str and desc_str not in unique_descriptions and desc_str != 'nan':
                            unique_descriptions.append(desc_str)
 
                    combined_description = " | ".join(unique_descriptions) if unique_descriptions else ""
 
                    billable_duration = day_df[day_df["Billable"] == "Yes"]["formatted_duration"].sum()
                    free_duration = day_df[day_df["Billable"] == "No"]["formatted_duration"].sum()
                    total_duration = day_df["formatted_duration"].sum()
 
                    summary_sheet.write(row, 0, sanitize_excel_cell(day_str), cell_format)
                    summary_sheet.write(row, 1, sanitize_excel_cell(combined_description), cell_wrap_format)
 
                    if format_choice == "hours":
                        summary_sheet.write_number(row, 2, billable_duration, time_format)
                        summary_sheet.write_number(row, 3, free_duration, time_format)
                        summary_sheet.write_number(row, 4, total_duration, time_format)
                    else:
                        summary_sheet.write_number(row, 2, billable_duration, number_format)
                        summary_sheet.write_number(row, 3, free_duration, number_format)
                        summary_sheet.write_number(row, 4, total_duration, number_format)
 
                    desc_length = len(combined_description)
                    lines_needed = max(1, (desc_length // 120) + 1)
                    summary_sheet.set_row(row, 18 * lines_needed)
 
                row += 1
 
            row += 1
 
            billable_df = user_df[user_df["Billable"] == "Yes"]
            non_billable_df = user_df[user_df["Billable"] == "No"]
 
            total_billable = billable_df["formatted_duration"].sum()
            total_non_billable = non_billable_df["formatted_duration"].sum()
            total_overall = user_df["formatted_duration"].sum()
 
            summary_sheet.merge_range(row, 0, row, 1, "BILLABLE TOTAL", green_format)
            if format_choice == "hours":
                summary_sheet.write_number(row, 2, total_billable, green_time_format)
            else:
                summary_sheet.write_number(row, 2, total_billable, green_number_format)
            summary_sheet.write(row, 3, "", green_format)
            summary_sheet.write(row, 4, "", green_format)
            summary_sheet.set_row(row, 20)
            row += 1
 
            summary_sheet.merge_range(row, 0, row, 1, "FREE TOTAL", red_format)
            summary_sheet.write(row, 2, "", red_format)
            if format_choice == "hours":
                summary_sheet.write_number(row, 3, total_non_billable, red_time_format)
            else:
                summary_sheet.write_number(row, 3, total_non_billable, red_number_format)
            summary_sheet.write(row, 4, "", red_format)
            summary_sheet.set_row(row, 20)
            row += 1
 
            summary_sheet.merge_range(row, 0, row, 1, "GRAND TOTAL", yellow_format)
            summary_sheet.write(row, 2, "", yellow_format)
            summary_sheet.write(row, 3, "", yellow_format)
            if format_choice == "hours":
                summary_sheet.write_number(row, 4, total_overall, yellow_time_format)
            else:
                summary_sheet.write_number(row, 4, total_overall, yellow_number_format)
            summary_sheet.set_row(row, 20)
            row += 3
 
        summary_sheet.set_column(0, 0, 14)
        summary_sheet.set_column(1, 1, 90)
        summary_sheet.set_column(2, 2, 15)
        summary_sheet.set_column(3, 3, 15)
        summary_sheet.set_column(4, 4, 15)
        summary_sheet.set_column(5, 5, 22)
 
        # ============== SAYFA 2: DETAYLI RAPOR ==============
        detail_sheet = workbook.add_worksheet("Detailed Report")
        detail_sheet.fit_to_pages(1, 0)
        detail_sheet.set_landscape()
        detail_sheet.set_paper(9)
 
        detail_row = 0
 
        detail_sheet.merge_range(detail_row, 0, detail_row, 4, "Detailed Time Report", header_format)
        detail_sheet.set_row(detail_row, 22)
        detail_row += 1
 
        if company_info:
            table_start_row = detail_row
 
            detail_sheet.write(detail_row, 0, "Company:", info_label_format)
            detail_sheet.merge_range(detail_row, 1, detail_row, 4, sanitize_excel_cell(company_info.get('company_name', '')), info_value_format)
            detail_sheet.set_row(detail_row, 18)
            detail_row += 1
 
            if company_info.get('contact_person'):
                detail_sheet.write(detail_row, 0, "Contact:", info_label_format)
                detail_sheet.merge_range(detail_row, 1, detail_row, 4, sanitize_excel_cell(company_info.get('contact_person', '')), info_value_format)
                detail_sheet.set_row(detail_row, 18)
                detail_row += 1
 
            if company_info.get('phone'):
                detail_sheet.write(detail_row, 0, "Phone:", info_label_format)
                detail_sheet.merge_range(detail_row, 1, detail_row, 4, sanitize_excel_cell(company_info.get('phone', '')), info_value_format)
                detail_sheet.set_row(detail_row, 18)
                detail_row += 1
 
            if company_info.get('address'):
                detail_sheet.write(detail_row, 0, "Address:", info_label_format)
                detail_sheet.merge_range(detail_row, 1, detail_row, 4, sanitize_excel_cell(company_info.get('address', '')), info_value_format)
                detail_sheet.set_row(detail_row, 18)
                detail_row += 1
 
            detail_row += 1
 
        detail_sheet.write(detail_row, 0, "Period:", info_label_format)
        detail_sheet.merge_range(detail_row, 1, detail_row, 4, sanitize_excel_cell(report_period), info_value_format)
        detail_sheet.set_row(detail_row, 18)
        detail_row += 1
 
        detail_sheet.write(detail_row, 0, "Projects:", info_label_format)
        detail_sheet.merge_range(detail_row, 1, detail_row, 4, sanitize_excel_cell(projects), info_value_format)
        detail_sheet.set_row(detail_row, 18)
        detail_row += 1
 
        detail_sheet.write(detail_row, 0, "Clients:", info_label_format)
        detail_sheet.merge_range(detail_row, 1, detail_row, 4, sanitize_excel_cell(customers), info_value_format)
        detail_sheet.set_row(detail_row, 18)
        detail_row += 2
 
        df_sorted = df.sort_values(['User', 'ParsedDate', 'Start Time'])
 
        for user_value in sorted(df_sorted['User'].dropna().unique()):
            user_df = df_sorted[df_sorted['User'] == user_value]
 
            detail_sheet.merge_range(detail_row, 0, detail_row, 4, sanitize_excel_cell(f"User: {user_value}"), user_header_format)
            detail_sheet.set_row(detail_row, 22)
            detail_row += 1
 
            for day_dt, day_full_str in zip(all_days, all_days_full):
                day_df = user_df[user_df['DayFull'] == day_full_str]
 
                if day_df.empty:
                    detail_sheet.merge_range(detail_row, 0, detail_row, 4, sanitize_excel_cell(f"Date: {day_full_str}"), detail_date_empty_format)
                    detail_sheet.set_row(detail_row, 20)
                    detail_row += 1
 
                    headers = ["Start Time", "End Time", "Duration", "Description", "Billable"]
                    for col_idx, header in enumerate(headers):
                        detail_sheet.write(detail_row, col_idx, header, detail_header_format)
                    detail_sheet.set_row(detail_row, 18)
                    detail_row += 1
 
                    detail_sheet.write(detail_row, 0, "-", detail_empty_cell)
                    detail_sheet.write(detail_row, 1, "-", detail_empty_cell)
                    detail_sheet.write(detail_row, 2, "-", detail_empty_cell)
                    detail_sheet.write(detail_row, 3, "No activity", detail_empty_wrap)
                    detail_sheet.write(detail_row, 4, "-", detail_empty_cell)
                    detail_sheet.set_row(detail_row, 16)
                    detail_row += 2
                else:
                    detail_sheet.merge_range(detail_row, 0, detail_row, 4, sanitize_excel_cell(f"Date: {day_full_str}"), detail_date_header_format)
                    detail_sheet.set_row(detail_row, 20)
                    detail_row += 1
 
                    headers = ["Start Time", "End Time", "Duration", "Description", "Billable"]
                    for col_idx, header in enumerate(headers):
                        detail_sheet.write(detail_row, col_idx, header, detail_header_format)
                    detail_sheet.set_row(detail_row, 18)
                    detail_row += 1
 
                    for idx, row_data in day_df.iterrows():
                        detail_sheet.write(detail_row, 0, sanitize_excel_cell(str(row_data.get('Start Time', ''))), detail_cell_center)
                        detail_sheet.write(detail_row, 1, sanitize_excel_cell(str(row_data.get('End Time', ''))), detail_cell_center)
 
                        if format_choice == "hours":
                            detail_sheet.write_number(detail_row, 2, row_data['formatted_duration'], detail_time_format)
                        else:
                            detail_sheet.write_number(detail_row, 2, row_data['formatted_duration'], detail_number_format)
 
                        desc_text = sanitize_excel_cell(str(row_data.get('Description', '')))
                        detail_sheet.write(detail_row, 3, desc_text, detail_cell_wrap)
                        detail_sheet.write(detail_row, 4, sanitize_excel_cell(str(row_data.get('Billable', 'No'))), detail_cell_center)
 
                        desc_length = len(desc_text)
                        lines_needed = max(1, (desc_length // 60) + 1)
                        detail_sheet.set_row(detail_row, 16 * lines_needed)
                        detail_row += 1
 
                    detail_row += 1
 
            detail_row += 1
 
        detail_sheet.set_column(0, 0, 10)
        detail_sheet.set_column(1, 1, 10)
        detail_sheet.set_column(2, 2, 10)
        detail_sheet.set_column(3, 3, 84)
        detail_sheet.set_column(4, 4, 8)
        detail_sheet.set_column(5, 5, 22)
 
    output.seek(0)
    return output

def generate_invoice_excel(data, logo_data=None, company_info=None):
    """Invoice Excel dosyası oluştur - A4 Landscape: Daha geniş, optimize fontlar"""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Side, Border
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter
    from PIL import Image as PILImage
    from io import BytesIO
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Tax Invoice"
    
    # A4 Landscape sayfa ayarları
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    
    # Kenar boşlukları - daha dar (daha fazla alan)
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.6
    ws.page_margins.bottom = 0.6
    
    # Formatlar - DAHA KÜÇÜK VE OKUNAKLI
    title_font = Font(bold=True, size=14, name='Calibri')
    bold_font = Font(bold=True, size=9.5, name='Calibri')
    regular_font = Font(size=9, name='Calibri')
    small_font = Font(size=8.5, name='Calibri')
    tiny_font = Font(size=8, name='Calibri')
    note_font = Font(size=7, color='7F8C8D', name='Calibri')
    company_name_font = Font(bold=True, size=10.5, name='Calibri')
    
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_center_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_center_alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
    left_top_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    border_side = Side(style='thin', color='000000')
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    thick_border_side = Side(style='medium', color='000000')
    
    # Sütun genişlikleri - DAHA GENİŞ (toplam ~100 birim)
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 26
    ws.column_dimensions['C'].width = 26
    ws.column_dimensions['D'].width = 12.5
    ws.column_dimensions['E'].width = 12.5
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 11
    
    row = 1
    
    # Başlık
    ws.cell(row=row, column=1, value="Tax Invoice")
    ws.cell(row=row, column=1).font = title_font
    ws.cell(row=row, column=1).alignment = center_alignment
    ws.merge_cells(f'A{row}:G{row}')
    ws.row_dimensions[row].height = 23
    row += 1
    
    # Alt başlık
    note_text = "(SUPPLY MEANT FOR EXPORT/ SUPPLY TO SEZ UNIT OR SEZ DEVELOPER FOR AUTHORISED OPERATIONS UNDER BOND OR LETTER OF UNDERTAKING WITHOUT PAYMENT OF IGST)"
    ws.cell(row=row, column=1, value=note_text)
    ws.cell(row=row, column=1).font = note_font
    ws.cell(row=row, column=1).alignment = center_alignment
    ws.merge_cells(f'A{row}:G{row}')
    ws.row_dimensions[row].height = 22
    row += 1
    
    company_start_row = row
    
    # ============== SOL TARAF ==============
    ws.merge_cells(f'A{row}:A{row+11}')
    ws.cell(row=row, column=1).border = border
    for i in range(12):
        ws.cell(row=row+i, column=1).border = border
    
    # Logo
    logo_start_row = row
    logo_end_row = row + 4
    
    if logo_data:
        try:
            if isinstance(logo_data, dict) and 'data' in logo_data:
                logo_bytes = logo_data['data']
            else:
                logo_bytes = logo_data
            
            temp_logo_io = BytesIO(logo_bytes)
            pil_image = PILImage.open(temp_logo_io)
            
            max_width = 350
            max_height = 85
            
            pil_image.thumbnail((max_width, max_height), PILImage.Resampling.LANCZOS)
            
            optimized_logo_io = BytesIO()
            pil_image.save(optimized_logo_io, format='PNG')
            optimized_logo_io.seek(0)
            
            logo = XLImage(optimized_logo_io)
            ws.add_image(logo, f'B{logo_start_row}')
            
        except Exception as e:
            pass
    
    ws.merge_cells(f'B{logo_start_row}:C{logo_end_row}')
    for i in range(logo_start_row, logo_end_row + 1):
        ws.cell(row=i, column=2).border = border
        ws.cell(row=i, column=3).border = border
        ws.cell(row=i, column=2).alignment = center_alignment
        ws.row_dimensions[i].height = 19
    
    row = logo_end_row + 1
    
    # Şirket bilgileri
    if company_info:
        company_name = company_info.get('company_name', 'ULEPUS')
        company_address = company_info.get('address', 'ODTÜ Teknokent Mustafa Kemal Mah. Dumlupınar Blv. No:280/G İç Kapı No:305 Çankaya/Ankara')
        company_phone = company_info.get('phone', '+90-312-486-1158')
        company_email = company_info.get('email', 'info@ulepus.com')
    else:
        company_name = 'ULEPUS'
        company_address = 'ODTÜ Teknokent Mustafa Kemal Mah. Dumlupınar Blv. No:280/G İç Kapı No:305 Çankaya/Ankara'
        company_phone = '+90-312-486-1158'
        company_email = 'info@ulepus.com'
    
    ws.cell(row=row, column=2, value=company_name)
    ws.cell(row=row, column=2).font = company_name_font
    ws.cell(row=row, column=2).alignment = left_center_alignment
    ws.cell(row=row, column=2).border = border
    ws.merge_cells(f'B{row}:C{row}')
    ws.row_dimensions[row].height = 19
    row += 1
    
    ws.cell(row=row, column=2, value=company_address)
    ws.cell(row=row, column=2).font = tiny_font
    ws.cell(row=row, column=2).alignment = left_top_alignment
    ws.cell(row=row, column=2).border = border
    ws.merge_cells(f'B{row}:C{row+2}')
    
    for i in range(3):
        for col in ['B', 'C']:
            ws[f'{col}{row+i}'].border = border
        ws.row_dimensions[row+i].height = 17
    
    row += 3
    
    ws.cell(row=row, column=2, value=company_email)
    ws.cell(row=row, column=2).font = tiny_font
    ws.cell(row=row, column=2).alignment = left_center_alignment
    ws.cell(row=row, column=2).border = border
    ws.merge_cells(f'B{row}:C{row}')
    ws.row_dimensions[row].height = 17
    row += 1
    
    ws.cell(row=row, column=2, value=f"Contact: {company_phone}")
    ws.cell(row=row, column=2).font = tiny_font
    ws.cell(row=row, column=2).alignment = left_center_alignment
    ws.cell(row=row, column=2).border = border
    ws.merge_cells(f'B{row}:C{row}')
    ws.row_dimensions[row].height = 17
    row += 1
    
    for c in [2, 3]:
        ws.cell(row=row, column=c).border = Border(
            bottom=thick_border_side, 
            left=border_side, 
            right=border_side, 
            top=border_side
        )
    ws.merge_cells(f'B{row}:C{row}')
    ws.row_dimensions[row].height = 7
    row += 1
    
    ws.cell(row=row, column=2, value="Buyer (Bill To):")
    ws.cell(row=row, column=2).font = bold_font
    ws.cell(row=row, column=2).alignment = left_center_alignment
    ws.cell(row=row, column=2).border = border
    ws.merge_cells(f'B{row}:C{row}')
    ws.row_dimensions[row].height = 17
    row += 1
    
    ws.cell(row=row, column=2, value=data.get('buyer_name', ''))
    ws.cell(row=row, column=2).font = bold_font
    ws.cell(row=row, column=2).alignment = left_center_alignment
    ws.cell(row=row, column=2).border = border
    ws.merge_cells(f'B{row}:C{row}')
    ws.row_dimensions[row].height = 17
    row += 1
    
    buyer_address = data.get('buyer_address', '')
    if buyer_address:
        ws.cell(row=row, column=2, value=buyer_address)
        ws.cell(row=row, column=2).font = tiny_font
        ws.cell(row=row, column=2).alignment = left_top_alignment
        ws.cell(row=row, column=2).border = border
        ws.merge_cells(f'B{row}:C{row+1}')
        
        for i in range(2):
            for col in ['B', 'C']:
                ws[f'{col}{row+i}'].border = border
            ws.row_dimensions[row+i].height = 17
        
        row += 2
    
    buyer_state = data.get('buyer_state', '')
    if buyer_state:
        ws.cell(row=row, column=2, value=f"State Name: {buyer_state}")
        ws.cell(row=row, column=2).font = tiny_font
        ws.cell(row=row, column=2).alignment = left_center_alignment
        ws.cell(row=row, column=2).border = border
        ws.merge_cells(f'B{row}:C{row}')
        ws.row_dimensions[row].height = 17
        row += 1
    
    place_of_supply = data.get('place_of_supply', '')
    if place_of_supply:
        ws.cell(row=row, column=2, value=f"Place of Supply: {place_of_supply}")
        ws.cell(row=row, column=2).font = tiny_font
        ws.cell(row=row, column=2).alignment = left_center_alignment
        ws.cell(row=row, column=2).border = border
        ws.merge_cells(f'B{row}:C{row}')
        ws.row_dimensions[row].height = 17
        row += 1
    
    contact_person = data.get('contact_person', '')
    if contact_person:
        ws.cell(row=row, column=2, value=f"Contact Person: {contact_person}")
        ws.cell(row=row, column=2).font = tiny_font
        ws.cell(row=row, column=2).alignment = left_center_alignment
        ws.cell(row=row, column=2).border = border
        ws.merge_cells(f'B{row}:C{row}')
        ws.row_dimensions[row].height = 17
        row += 1
    
    buyer_email = data.get('buyer_email', '')
    if buyer_email:
        ws.cell(row=row, column=2, value=f"E-Mail: {buyer_email}")
        ws.cell(row=row, column=2).font = tiny_font
        ws.cell(row=row, column=2).alignment = left_center_alignment
        ws.cell(row=row, column=2).border = border
        ws.merge_cells(f'B{row}:C{row}')
        ws.row_dimensions[row].height = 17
        row += 1
    
    # ============== SAĞ TARAF ==============
    detail_row = company_start_row
    
    details = [
        ("Invoice No.:", data.get('invoice_no', ''), "Dated:", data.get('invoice_date', '')),
        ("Delivery Note:", data.get('delivery_note', ''), "Payment Terms:", data.get('payment_terms', '')),
        ("Ref. & Date:", data.get('ref_date', ''), "Other Ref:", data.get('other_references', '')),
        ("Buyer Order No:", data.get('buyer_order_no', ''), "Dated:", data.get('order_date', '')),
        ("Dispatch Doc:", data.get('dispatch_doc_no', ''), "Del. Note Date:", data.get('delivery_note_date', '')),
        ("Dispatched:", data.get('dispatched_through', ''), "Destination:", data.get('destination', '')),
    ]
    
    for label1, value1, label2, value2 in details:
        ws.cell(row=detail_row, column=4, value=label1)
        ws.cell(row=detail_row, column=4).font = Font(bold=True, size=8.5, name='Calibri')
        ws.cell(row=detail_row, column=4).alignment = center_alignment
        ws.cell(row=detail_row, column=4).border = border
        ws.merge_cells(f'D{detail_row}:E{detail_row}')
        ws.row_dimensions[detail_row].height = 15
        
        ws.cell(row=detail_row + 1, column=4, value=value1)
        ws.cell(row=detail_row + 1, column=4).font = tiny_font
        ws.cell(row=detail_row + 1, column=4).alignment = center_alignment
        ws.cell(row=detail_row + 1, column=4).border = border
        ws.merge_cells(f'D{detail_row + 1}:E{detail_row + 1}')
        ws.row_dimensions[detail_row + 1].height = 15
        
        ws.cell(row=detail_row, column=6, value=label2)
        ws.cell(row=detail_row, column=6).font = Font(bold=True, size=8.5, name='Calibri')
        ws.cell(row=detail_row, column=6).alignment = center_alignment
        ws.cell(row=detail_row, column=6).border = border
        ws.merge_cells(f'F{detail_row}:G{detail_row}')
        
        ws.cell(row=detail_row + 1, column=6, value=value2)
        ws.cell(row=detail_row + 1, column=6).font = tiny_font
        ws.cell(row=detail_row + 1, column=6).alignment = center_alignment
        ws.cell(row=detail_row + 1, column=6).border = border
        ws.merge_cells(f'F{detail_row + 1}:G{detail_row + 1}')
        
        detail_row += 2
    
    single_details = [
        ("LUT/Bond No.:", data.get('lut_bond_no', '')),
        ("Country:", data.get('country', '')),
    ]
    
    for label, value in single_details:
        ws.cell(row=detail_row, column=4, value=label)
        ws.cell(row=detail_row, column=4).font = Font(bold=True, size=8.5, name='Calibri')
        ws.cell(row=detail_row, column=4).alignment = center_alignment
        ws.cell(row=detail_row, column=4).border = border
        ws.merge_cells(f'D{detail_row}:G{detail_row}')
        ws.row_dimensions[detail_row].height = 15
        
        ws.cell(row=detail_row + 1, column=4, value=value)
        ws.cell(row=detail_row + 1, column=4).font = tiny_font
        ws.cell(row=detail_row + 1, column=4).alignment = center_alignment
        ws.cell(row=detail_row + 1, column=4).border = border
        ws.merge_cells(f'D{detail_row + 1}:G{detail_row + 1}')
        ws.row_dimensions[detail_row + 1].height = 15
        
        detail_row += 2
    
    ws.cell(row=detail_row, column=4, value="From:")
    ws.cell(row=detail_row, column=4).font = Font(bold=True, size=8.5, name='Calibri')
    ws.cell(row=detail_row, column=4).alignment = center_alignment
    ws.cell(row=detail_row, column=4).border = border
    ws.merge_cells(f'D{detail_row}:E{detail_row}')
    ws.row_dimensions[detail_row].height = 15
    
    ws.cell(row=detail_row + 1, column=4, value=data.get('from', ''))
    ws.cell(row=detail_row + 1, column=4).font = tiny_font
    ws.cell(row=detail_row + 1, column=4).alignment = center_alignment
    ws.cell(row=detail_row + 1, column=4).border = border
    ws.merge_cells(f'D{detail_row + 1}:E{detail_row + 1}')
    ws.row_dimensions[detail_row + 1].height = 15
    
    ws.cell(row=detail_row, column=6, value="To:")
    ws.cell(row=detail_row, column=6).font = Font(bold=True, size=8.5, name='Calibri')
    ws.cell(row=detail_row, column=6).alignment = center_alignment
    ws.cell(row=detail_row, column=6).border = border
    ws.merge_cells(f'F{detail_row}:G{detail_row}')
    
    ws.cell(row=detail_row + 1, column=6, value=data.get('to', ''))
    ws.cell(row=detail_row + 1, column=6).font = tiny_font
    ws.cell(row=detail_row + 1, column=6).alignment = center_alignment
    ws.cell(row=detail_row + 1, column=6).border = border
    ws.merge_cells(f'F{detail_row + 1}:G{detail_row + 1}')
    
    detail_row += 2
    
    ws.cell(row=detail_row, column=4, value="Terms of Delivery:")
    ws.cell(row=detail_row, column=4).font = Font(bold=True, size=8.5, name='Calibri')
    ws.cell(row=detail_row, column=4).alignment = center_alignment
    ws.cell(row=detail_row, column=4).border = border
    ws.merge_cells(f'D{detail_row}:G{detail_row}')
    ws.row_dimensions[detail_row].height = 15
    
    ws.cell(row=detail_row + 1, column=4, value=data.get('terms_of_delivery', ''))
    ws.cell(row=detail_row + 1, column=4).font = tiny_font
    ws.cell(row=detail_row + 1, column=4).alignment = center_alignment
    ws.cell(row=detail_row + 1, column=4).border = border
    ws.merge_cells(f'D{detail_row + 1}:G{detail_row + 1}')
    ws.row_dimensions[detail_row + 1].height = 15
    
    # ============== HİZMETLER ==============
    row = max(row, detail_row + 2)
    
    headers = ["Sl No.", "Description of Goods", "HSN/SAC", "Quantity", "Rate", "Per", "Amount"]
    header_row = row
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = header
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.fill = header_fill
        cell.border = border
    
    ws.row_dimensions[header_row].height = 20
    row += 1
    
    services = data.get('services', [])
    total_quantity = 0
    total_amount = 0
    
    for idx, service in enumerate(services, start=1):
        quantity = float(service.get('quantity', 0))
        rate = float(service.get('rate', 0))
        amount = quantity * rate
        total_quantity += quantity
        total_amount += amount
        
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=1).font = regular_font
        ws.cell(row=row, column=1).alignment = center_alignment
        ws.cell(row=row, column=1).border = border
        
        ws.cell(row=row, column=2, value=service.get('description', ''))
        ws.cell(row=row, column=2).font = regular_font
        ws.cell(row=row, column=2).alignment = left_center_alignment
        ws.cell(row=row, column=2).border = border
        
        ws.cell(row=row, column=3, value=service.get('hsn', ''))
        ws.cell(row=row, column=3).font = regular_font
        ws.cell(row=row, column=3).alignment = center_alignment
        ws.cell(row=row, column=3).border = border
        
        ws.cell(row=row, column=4, value=quantity)
        ws.cell(row=row, column=4).font = regular_font
        ws.cell(row=row, column=4).alignment = center_alignment
        ws.cell(row=row, column=4).border = border
        
        ws.cell(row=row, column=5, value=rate)
        ws.cell(row=row, column=5).font = regular_font
        ws.cell(row=row, column=5).alignment = right_center_alignment
        ws.cell(row=row, column=5).border = border
        ws.cell(row=row, column=5).number_format = '#,##0.00'
        
        ws.cell(row=row, column=6, value=service.get('per', ''))
        ws.cell(row=row, column=6).font = regular_font
        ws.cell(row=row, column=6).alignment = center_alignment
        ws.cell(row=row, column=6).border = border
        
        ws.cell(row=row, column=7, value=amount)
        ws.cell(row=row, column=7).font = regular_font
        ws.cell(row=row, column=7).alignment = right_center_alignment
        ws.cell(row=row, column=7).border = border
        ws.cell(row=row, column=7).number_format = '#,##0.00'
        
        ws.row_dimensions[row].height = 18
        row += 1
    
    ws.cell(row=row, column=1, value="Total")
    ws.cell(row=row, column=1).font = bold_font
    ws.cell(row=row, column=1).alignment = center_alignment
    ws.cell(row=row, column=1).border = border
    ws.merge_cells(f'A{row}:B{row}')
    
    ws.cell(row=row, column=3, value="")
    ws.cell(row=row, column=3).border = border
    
    ws.cell(row=row, column=4, value=total_quantity)
    ws.cell(row=row, column=4).font = bold_font
    ws.cell(row=row, column=4).alignment = center_alignment
    ws.cell(row=row, column=4).border = border
    
    ws.cell(row=row, column=5, value="")
    ws.cell(row=row, column=5).border = border
    
    ws.cell(row=row, column=6, value="")
    ws.cell(row=row, column=6).border = border
    
    ws.cell(row=row, column=7, value=total_amount)
    ws.cell(row=row, column=7).font = bold_font
    ws.cell(row=row, column=7).alignment = right_center_alignment
    ws.cell(row=row, column=7).border = border
    ws.cell(row=row, column=7).number_format = '#,##0.00'
    
    ws.row_dimensions[row].height = 20
    row += 1
    
    try:
        from num2words import num2words
        amount_words = num2words(int(total_amount), lang='en').title() + " Euro Only"
    except:
        amount_words = f"{int(total_amount)} Euro Only"
    
    ws.cell(row=row, column=1, value=f"Amount Chargeable (in words): {amount_words}")
    ws.cell(row=row, column=1).font = regular_font
    ws.cell(row=row, column=1).alignment = left_center_alignment
    ws.cell(row=row, column=1).border = border
    ws.merge_cells(f'A{row}:G{row}')
    ws.row_dimensions[row].height = 18
    row += 1
    
    ws.cell(row=row, column=1, value="E. & O.E")
    ws.cell(row=row, column=1).font = bold_font
    ws.cell(row=row, column=1).alignment = left_center_alignment
    ws.cell(row=row, column=1).border = border
    ws.merge_cells(f'A{row}:G{row}')
    ws.row_dimensions[row].height = 16
    row += 1
    
    # ============== DECLARATION & BANK ==============
    declaration_start_row = row
    
    ws.cell(row=row, column=1, value="Declaration")
    ws.cell(row=row, column=1).font = bold_font
    ws.cell(row=row, column=1).alignment = left_center_alignment
    ws.cell(row=row, column=1).border = border
    ws.merge_cells(f'A{row}:C{row}')
    ws.row_dimensions[row].height = 18
    
    ws.cell(row=row, column=4, value="Company's Bank Details")
    ws.cell(row=row, column=4).font = bold_font
    ws.cell(row=row, column=4).alignment = left_center_alignment
    ws.cell(row=row, column=4).border = border
    ws.merge_cells(f'D{row}:G{row}')
    
    row += 1
    
    ws.cell(row=row, column=1, value="Terms & Conditions:")
    ws.cell(row=row, column=1).font = Font(bold=True, size=8.5, name='Calibri')
    ws.cell(row=row, column=1).alignment = left_center_alignment
    ws.cell(row=row, column=1).border = border
    ws.merge_cells(f'A{row}:C{row}')
    ws.row_dimensions[row].height = 16
    row += 1
    
    terms = [
        "1. All layout services will be carried out from the ULEPUS office.",
        "2. The customer will provide all the software licenses required, accessed through a secure VPN tunnel.",
        "3. The customer will retain responsibility for the circuit design at all times.",
        "4. Rates apply to a 40 hour working week."
    ]
    
    bank_details = [
        f"A/c Holder's Name: {data.get('bank_holder', '')}",
        f"Bank Name: {data.get('bank_name', '')}",
        f"A/c No.: {data.get('bank_account', '')}",
        f"Branch & IFS Code: {data.get('bank_branch', '')}",
        f"SWIFT Code: {data.get('bank_swift', '')}"
    ]
    
    bank_row = declaration_start_row + 1
    for i, term in enumerate(terms):
        ws.cell(row=row, column=1, value=term)
        ws.cell(row=row, column=1).font = tiny_font
        ws.cell(row=row, column=1).alignment = left_top_alignment
        ws.cell(row=row, column=1).border = border
        ws.merge_cells(f'A{row}:C{row}')
        ws.row_dimensions[row].height = 18
        
        if i < len(bank_details):
            ws.cell(row=bank_row, column=4, value=bank_details[i])
            ws.cell(row=bank_row, column=4).font = tiny_font
            ws.cell(row=bank_row, column=4).alignment = left_center_alignment
            ws.cell(row=bank_row, column=4).border = border
            ws.merge_cells(f'D{bank_row}:G{bank_row}')
            ws.row_dimensions[bank_row].height = 18
            bank_row += 1
        
        row += 1
    
    if len(bank_details) > len(terms):
        ws.cell(row=bank_row, column=4, value=bank_details[4])
        ws.cell(row=bank_row, column=4).font = tiny_font
        ws.cell(row=bank_row, column=4).alignment = left_center_alignment
        ws.cell(row=bank_row, column=4).border = border
        ws.merge_cells(f'D{bank_row}:G{bank_row}')
        ws.row_dimensions[bank_row].height = 18
        
        ws.cell(row=row, column=1, value="")
        ws.cell(row=row, column=1).border = border
        ws.merge_cells(f'A{row}:C{row}')
        ws.row_dimensions[row].height = 18
        row += 1
    
    final_row = row - 1
    
    # ============== BORDER YÖNETİMİ ==============
    
    for c in range(1, 8):
        ws.cell(row=1, column=c).border = Border(
            top=thick_border_side,
            left=thick_border_side if c == 1 else border_side,
            right=thick_border_side if c == 7 else border_side,
            bottom=border_side
        )
    
    for r in range(2, final_row):
        current_left = ws.cell(row=r, column=1).border
        ws.cell(row=r, column=1).border = Border(
            left=thick_border_side,
            right=current_left.right if current_left else border_side,
            top=current_left.top if current_left else border_side,
            bottom=current_left.bottom if current_left else border_side
        )
        
        current_right = ws.cell(row=r, column=7).border
        ws.cell(row=r, column=7).border = Border(
            right=thick_border_side,
            left=current_right.left if current_right else border_side,
            top=current_right.top if current_right else border_side,
            bottom=current_right.bottom if current_right else border_side
        )
    
    for c in range(1, 8):
        current_bottom = ws.cell(row=final_row, column=c).border
        ws.cell(row=final_row, column=c).border = Border(
            left=thick_border_side if c == 1 else (current_bottom.left if current_bottom else border_side),
            right=thick_border_side if c == 7 else (current_bottom.right if current_bottom else border_side),
            top=current_bottom.top if current_bottom else border_side,
            bottom=thick_border_side
        )
    
    for r in range(1, final_row + 1):
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            if not cell.border or cell.border == Border():
                cell.border = border
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

# ============== AUTH ENDPOINTS ==============

@app.route('/api/auth/register', methods=['POST'])
@limiter.limit("5 per minute")
def register():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
 
        email = data.get('email', '').strip().lower()
        valid, message = validate_email(email)
        if not valid:
            return jsonify({'error': message}), 400
 
        if mongo.db.users.find_one({'email': email}):
            return jsonify({'error': 'Email already registered'}), 400
 
        password = data.get('password', '')
        valid, message = validate_password(password)
        if not valid:
            return jsonify({'error': message}), 400
 
        user_type = data.get('user_type', 'individual')
        if user_type not in ['individual', 'company', 'employee']:
            return jsonify({'error': 'Invalid user type'}), 400
 
        data_source = data.get('data_source', 'csv')
        if data_source not in ['csv', 'clockify']:
            return jsonify({'error': 'Invalid data source'}), 400
 
        user_doc = {
            'email': email,
            'password_hash': generate_password_hash(password),
            'user_type': user_type,
            'data_source': data_source,
            'created_at': datetime.utcnow(),
            'last_login': None
        }
 
        # --- EMPLOYEE kaydı ---
        if user_type == 'employee':
            company_code = data.get('company_code', '').strip()
            if not company_code:
                return jsonify({'error': 'Company code is required for employee registration'}), 400
 
            # Şirketi bul
            company = mongo.db.users.find_one({
                'user_type': 'company',
                'company_profile.company_code': company_code
            })
            if not company:
                return jsonify({'error': 'Invalid company code. Please check with your employer.'}), 404
 
            full_name = sanitize_input(data.get('full_name', ''), 100)
            if not full_name:
                return jsonify({'error': 'Full name is required'}), 400
 
            clockify_username = sanitize_input(data.get('clockify_username', ''), 100)
 
            user_doc['employee_profile'] = {
                'full_name': full_name,
                'phone': sanitize_input(data.get('phone', ''), 20),
                'clockify_username': clockify_username,
                'company_id': str(company['_id']),
                'company_name': company.get('company_profile', {}).get('company_name', ''),
                'company_code': company_code,
                'status': 'active'
            }
 
            # Clockify kullanılıyorsa, şirketin API key'ini kopyala
            if data_source == 'clockify' and company.get('clockify_api_key'):
                user_doc['clockify_api_key'] = company['clockify_api_key']
                user_doc['data_source'] = 'clockify'
            else:
                user_doc['data_source'] = company.get('data_source', 'csv')
 
        # --- INDIVIDUAL kaydı ---
        elif user_type == 'individual':
            full_name = sanitize_input(data.get('full_name', ''), 100)
            if not full_name:
                return jsonify({'error': 'Full name is required'}), 400
            user_doc['individual_profile'] = {
                'full_name': full_name,
                'phone': sanitize_input(data.get('phone', ''), 20)
            }
 
        # --- COMPANY kaydı ---
        else:
            company_name = sanitize_input(data.get('company_name', ''), 100)
            if not company_name:
                return jsonify({'error': 'Company name is required'}), 400
 
            # Benzersiz şirket kodu üret
            import secrets, string
            code_chars = string.ascii_uppercase + string.digits
            company_code = ''.join(secrets.choice(code_chars) for _ in range(8))
            # Çakışma kontrolü
            while mongo.db.users.find_one({'company_profile.company_code': company_code}):
                company_code = ''.join(secrets.choice(code_chars) for _ in range(8))
 
            company_profile = {
                'company_name': company_name,
                'company_code': company_code,
                'contact_person': sanitize_input(data.get('contact_person', ''), 100),
                'phone': sanitize_input(data.get('phone', ''), 20),
                'address': sanitize_input(data.get('address', ''), 500)
            }
 
            if 'logo_base64' in data and data['logo_base64']:
                try:
                    logo_data = base64.b64decode(data['logo_base64'].split(',')[1])
                    if len(logo_data) > 2 * 1024 * 1024:
                        return jsonify({'error': 'Logo size must be less than 2MB'}), 400
                    company_profile['logo_data'] = logo_data
                    company_profile['logo_mimetype'] = data.get('logo_mimetype', 'image/png')
                except Exception:
                    return jsonify({'error': 'Invalid logo data'}), 400
 
            user_doc['company_profile'] = company_profile
 
        if data_source == 'clockify' and data.get('clockify_api_key') and user_type != 'employee':
            api_key = data['clockify_api_key'].strip()
            if api_key:
                encrypted_key = encrypt_api_key(api_key)
                if encrypted_key:
                    user_doc['clockify_api_key'] = encrypted_key
                else:
                    return jsonify({'error': 'Failed to encrypt API key'}), 500
 
        result = mongo.db.users.insert_one(user_doc)
 
        response_data = {
            'message': 'User registered successfully',
            'user_id': str(result.inserted_id)
        }
        # Şirket kaydıysa kodu döndür
        if user_type == 'company':
            response_data['company_code'] = company_code
 
        return jsonify(response_data), 201
 
    except Exception as e:
        return safe_error_response(e, 500)

@app.route('/api/auth/login', methods=['POST'])
@limiter.limit("5 per minute")
@limiter.limit("20 per hour")
def login():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
 
        email = data.get('email', '').strip().lower()
        password = data.get('password', '')
 
        if not email or not password:
            return jsonify({'error': 'Email and password required'}), 400
 
        valid, message = validate_email(email)
        if not valid:
            return jsonify({'error': 'Invalid credentials'}), 401
 
        user = mongo.db.users.find_one({'email': email})
        if not user or not check_password_hash(user['password_hash'], password):
            return jsonify({'error': 'Invalid credentials'}), 401
 
        mongo.db.users.update_one(
            {'_id': user['_id']},
            {'$set': {'last_login': datetime.utcnow()}}
        )
 
        access_token = create_access_token(identity=str(user['_id']))
 
        user_info = {
            'id': str(user['_id']),
            'email': user['email'],
            'user_type': user['user_type'],
            'data_source': user.get('data_source', 'csv')
        }
 
        if user['user_type'] == 'individual':
            user_info['profile'] = user.get('individual_profile', {})
        elif user['user_type'] == 'employee':
            user_info['profile'] = user.get('employee_profile', {})
        else:
            profile = user.get('company_profile', {})
            if 'logo_data' in profile:
                logo_base64 = base64.b64encode(profile['logo_data']).decode('utf-8')
                profile = dict(profile)
                profile['logo_base64'] = f"data:{profile.get('logo_mimetype', 'image/png')};base64,{logo_base64}"
                del profile['logo_data']
            user_info['profile'] = profile
 
        return jsonify({'access_token': access_token, 'user': user_info}), 200
 
    except Exception as e:
        return safe_error_response(e, 500)

@app.route('/api/auth/me', methods=['GET'])
@jwt_required()
def get_current_user():
    try:
        user_id = get_jwt_identity()
        try:
            user_obj_id = ObjectId(user_id)
        except:
            return jsonify({'error': 'Invalid user ID'}), 401
 
        user = mongo.db.users.find_one({'_id': user_obj_id})
        if not user:
            return jsonify({'error': 'User not found'}), 404
 
        user_info = {
            'id': str(user['_id']),
            'email': user['email'],
            'user_type': user['user_type'],
            'data_source': user.get('data_source', 'csv')
        }
 
        if user['user_type'] == 'individual':
            user_info['profile'] = user.get('individual_profile', {})
        elif user['user_type'] == 'employee':
            user_info['profile'] = user.get('employee_profile', {})
        else:
            profile = user.get('company_profile', {})
            if 'logo_data' in profile:
                logo_base64 = base64.b64encode(profile['logo_data']).decode('utf-8')
                profile = dict(profile)
                profile['logo_base64'] = f"data:{profile.get('logo_mimetype', 'image/png')};base64,{logo_base64}"
                del profile['logo_data']
            user_info['profile'] = profile
 
        return jsonify(user_info), 200
 
    except Exception as e:
        return safe_error_response(e, 500)


@app.route('/api/company/info', methods=['GET'])
@jwt_required()
def get_company_info():
    """Çalışanın bağlı olduğu şirket bilgilerini döndür"""
    try:
        user_id = get_jwt_identity()
        user = mongo.db.users.find_one({'_id': ObjectId(user_id)})
 
        if not user or user.get('user_type') != 'employee':
            return jsonify({'error': 'Only employees can access this endpoint'}), 403
 
        company_id = user.get('employee_profile', {}).get('company_id')
        if not company_id:
            return jsonify({'error': 'No company linked'}), 404
 
        company = mongo.db.users.find_one({'_id': ObjectId(company_id)})
        if not company:
            return jsonify({'error': 'Company not found'}), 404
 
        profile = company.get('company_profile', {})
        company_info = {
            'company_name': profile.get('company_name', ''),
            'contact_person': profile.get('contact_person', ''),
            'phone': profile.get('phone', ''),
        }
        if 'logo_data' in profile:
            logo_b64 = base64.b64encode(profile['logo_data']).decode('utf-8')
            company_info['logo_base64'] = f"data:{profile.get('logo_mimetype','image/png')};base64,{logo_b64}"
 
        return jsonify(company_info), 200
 
    except Exception as e:
        return safe_error_response(e, 500)

@app.route('/api/company/employees', methods=['GET'])
@jwt_required()
def get_employees():
    """Şirkete bağlı tüm çalışanları listele (sadece company rolü)"""
    try:
        user_id = get_jwt_identity()
        user = mongo.db.users.find_one({'_id': ObjectId(user_id)})
 
        if not user or user.get('user_type') != 'company':
            return jsonify({'error': 'Only company accounts can access this endpoint'}), 403
 
        employees = list(mongo.db.users.find(
            {'user_type': 'employee', 'employee_profile.company_id': str(user['_id'])},
            {'password_hash': 0, 'clockify_api_key': 0}
        ))
 
        result = []
        for emp in employees:
            profile = emp.get('employee_profile', {})
            result.append({
                'id': str(emp['_id']),
                'email': emp['email'],
                'full_name': profile.get('full_name', ''),
                'phone': profile.get('phone', ''),
                'clockify_username': profile.get('clockify_username', ''),
                'status': profile.get('status', 'active'),
                'created_at': emp.get('created_at', '').isoformat() if emp.get('created_at') else ''
            })
 
        return jsonify(result), 200
 
    except Exception as e:
        return safe_error_response(e, 500)

@app.route('/api/company/employees/<employee_id>', methods=['PUT'])
@jwt_required()
def update_employee(employee_id):
    """Çalışan bilgilerini güncelle (sadece company rolü)"""
    try:
        user_id = get_jwt_identity()
        user = mongo.db.users.find_one({'_id': ObjectId(user_id)})
 
        if not user or user.get('user_type') != 'company':
            return jsonify({'error': 'Only company accounts can access this endpoint'}), 403
 
        data = request.get_json()
 
        employee = mongo.db.users.find_one({
            '_id': ObjectId(employee_id),
            'user_type': 'employee',
            'employee_profile.company_id': str(user['_id'])
        })
        if not employee:
            return jsonify({'error': 'Employee not found'}), 404
 
        update_data = {}
        if 'full_name' in data:
            update_data['employee_profile.full_name'] = sanitize_input(data['full_name'], 100)
        if 'phone' in data:
            update_data['employee_profile.phone'] = sanitize_input(data['phone'], 20)
        if 'clockify_username' in data:
            update_data['employee_profile.clockify_username'] = sanitize_input(data['clockify_username'], 100)
        if 'status' in data and data['status'] in ['active', 'inactive']:
            update_data['employee_profile.status'] = data['status']
 
        if update_data:
            mongo.db.users.update_one({'_id': ObjectId(employee_id)}, {'$set': update_data})
 
        return jsonify({'message': 'Employee updated successfully'}), 200
 
    except Exception as e:
        return safe_error_response(e, 500)
    
@app.route('/api/company/employees/<employee_id>', methods=['DELETE'])
@jwt_required()
def delete_employee(employee_id):
    """Çalışanı sistemden sil (sadece company rolü)"""
    try:
        user_id = get_jwt_identity()
        user = mongo.db.users.find_one({'_id': ObjectId(user_id)})
 
        if not user or user.get('user_type') != 'company':
            return jsonify({'error': 'Only company accounts can access this endpoint'}), 403
 
        result = mongo.db.users.delete_one({
            '_id': ObjectId(employee_id),
            'user_type': 'employee',
            'employee_profile.company_id': str(user['_id'])
        })
 
        if result.deleted_count == 0:
            return jsonify({'error': 'Employee not found'}), 404
 
        return jsonify({'message': 'Employee removed successfully'}), 200
 
    except Exception as e:
        return safe_error_response(e, 500)

@app.route('/api/company/code', methods=['GET'])
@jwt_required()
def get_company_code():
    """Şirket kodunu döndür (sadece company rolü)"""
    try:
        user_id = get_jwt_identity()
        user = mongo.db.users.find_one({'_id': ObjectId(user_id)})
 
        if not user or user.get('user_type') != 'company':
            return jsonify({'error': 'Only company accounts can access this endpoint'}), 403
 
        company_code = user.get('company_profile', {}).get('company_code', '')
        return jsonify({'company_code': company_code}), 200
 
    except Exception as e:
        return safe_error_response(e, 500)

@app.route('/api/clockify/employee-report', methods=['POST'])
@jwt_required()
def get_employee_clockify_report():
    """
    Çalışan kendi raporunu çıkarır.
    Şirketin Clockify API key'ini kullanır,
    raporu sadece kendi clockify_username'iyle filtreler.
    """
    try:
        user_id = get_jwt_identity()
        user = mongo.db.users.find_one({'_id': ObjectId(user_id)})
 
        if not user or user.get('user_type') != 'employee':
            return jsonify({'error': 'Only employees can access this endpoint'}), 403
 
        profile = user.get('employee_profile', {})
        clockify_username = profile.get('clockify_username', '')
        if not clockify_username:
            return jsonify({'error': 'Your Clockify username is not configured. Please contact your employer.'}), 400
 
        company_id = profile.get('company_id')
        if not company_id:
            return jsonify({'error': 'No company linked to your account'}), 400
 
        company = mongo.db.users.find_one({'_id': ObjectId(company_id)})
        if not company:
            return jsonify({'error': 'Company not found'}), 404
 
        encrypted_key = company.get('clockify_api_key', '')
        api_key = decrypt_api_key(encrypted_key) if encrypted_key else None
        if not api_key:
            return jsonify({'error': 'Company Clockify API key not configured. Please contact your employer.'}), 400
 
        data = request.get_json()
        workspace_id  = data.get('workspace_id')
        start_date    = data.get('start_date')
        end_date      = data.get('end_date')
        project_ids   = data.get('project_ids', [])
        format_choice = data.get('format', 'decimal')
 
        if not all([workspace_id, start_date, end_date]):
            return jsonify({'error': 'Missing required parameters'}), 400
 
        headers = {'X-Api-Key': api_key, 'Content-Type': 'application/json'}
 
        report_url = f'https://reports.api.clockify.me/v1/workspaces/{workspace_id}/reports/detailed'
        report_payload = {
            "dateRangeStart": start_date,
            "dateRangeEnd":   end_date,
            "detailedFilter": {"page": 1, "pageSize": 1000}
        }
        if project_ids:
            report_payload["detailedFilter"]["projects"] = {"ids": project_ids, "contains": "CONTAINS"}
 
        report_response = requests.post(report_url, headers=headers, json=report_payload, timeout=30)
        if report_response.status_code != 200:
            return jsonify({'error': f'Clockify API error: {report_response.text}'}), 400
 
        report_data  = report_response.json()
        time_entries = report_data.get('timeentries', [])
 
        # Sadece bu çalışana ait girdileri filtrele
        filtered_entries = [
            e for e in time_entries
            if (_safe_str(e.get('userName'), '')).lower() == clockify_username.lower()
        ]
 
        if not filtered_entries:
            return jsonify({'error': f'No time entries found for user "{clockify_username}" in the selected period.'}), 400
 
        csv_data = []
        for entry in filtered_entries:
            try:
                time_interval    = entry.get('timeInterval', {})
                start_str        = time_interval.get('start')
                end_str          = time_interval.get('end')
                duration_seconds = time_interval.get('duration', 0)
 
                if not start_str:
                    continue
 
                start_time = datetime.fromisoformat(start_str.replace('Z', '+00:00'))
                end_time   = datetime.fromisoformat(end_str.replace('Z', '+00:00')) if end_str else start_time + timedelta(seconds=duration_seconds)
                total_seconds = duration_seconds if duration_seconds > 0 else (end_time - start_time).total_seconds()
 
                h = int(total_seconds // 3600)
                m = int((total_seconds % 3600) // 60)
                s = int(total_seconds % 60)
 
                csv_data.append({
                    'Project':      _safe_str(entry.get('projectName'), 'No Project'),
                    'Client':       _safe_str(entry.get('clientName'),  'No Client'),
                    'User':         _safe_str(entry.get('userName'),    ''),
                    'Description':  entry.get('description', '') or '',
                    'Start Date':   start_time.strftime('%d/%m/%Y'),
                    'Start Time':   start_time.strftime('%H:%M:%S'),
                    'End Time':     end_time.strftime('%H:%M:%S'),
                    'Duration (h)': f"{h:02d}:{m:02d}:{s:02d}",
                    'Billable':     'Yes' if entry.get('billable', False) else 'No'
                })
            except Exception as ex:
                app.logger.warning(f"Error processing entry: {str(ex)}")
                continue
 
        if not csv_data:
            return jsonify({'error': 'No valid time entries found'}), 400
 
        df = pd.DataFrame(csv_data)
 
        # DataFrame sanitization
        df["Project"]      = df["Project"].apply(lambda x: _safe_str(x, "No Project"))
        df["Client"]       = df["Client"].apply(lambda x: _safe_str(x, "No Client"))
        df["Description"]  = df["Description"].fillna("").astype(str)
        df["Billable"]     = df["Billable"].fillna("No").astype(str)
        df["Duration (h)"] = df["Duration (h)"].fillna("00:00:00").astype(str)
 
        # ── DÜZELTİLMİŞ: sanitization SONRASI hesapla,
        #    "No Project" / "No Client" placeholder'larını filtrele ──
        real_projects  = [p for p in df["Project"].dropna().unique() if p != "No Project"]
        real_customers = [c for c in df["Client"].dropna().unique()  if c != "No Client"]
 
        overall_projects  = ", ".join(real_projects)  if real_projects  else "No Project"
        overall_customers = ", ".join(real_customers) if real_customers else "No Client"
 
        df['ParsedDate'] = pd.to_datetime(df['Start Date'], format='%d/%m/%Y', errors='coerce')
 
        company_profile = company.get('company_profile', {})
        logo_data    = None
        company_info = None
        if 'logo_data' in company_profile:
            logo_data = {'data': company_profile['logo_data'], 'mimetype': company_profile.get('logo_mimetype', 'image/png')}
        company_info = {
            'company_name':   company_profile.get('company_name', ''),
            'contact_person': company_profile.get('contact_person', ''),
            'phone':          company_profile.get('phone', ''),
            'address':        company_profile.get('address', '')
        }
 
        try:
            req_start = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
            req_end   = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
            if req_start.month == req_end.month and req_start.year == req_end.year:
                report_period = req_start.strftime("%B %Y")
            else:
                report_period = f"{req_start.strftime('%B %Y')} - {req_end.strftime('%B %Y')}"
        except:
            report_period = "All Data"
 
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output = generate_excel_report(
            df, format_choice, report_period,
            overall_projects, overall_customers,
            logo_data, company_info,
            date_range_start=start_date,
            date_range_end=end_date
        )
        filename = f"Report_{clockify_username}_{timestamp}.xlsx"
 
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
 
    except Exception as e:
        app.logger.error(f"Error: {str(e)}\n{traceback.format_exc()}")
        return jsonify({'error': str(e)}), 500


# ============== PROFILE ENDPOINTS ==============

@app.route('/api/profile', methods=['PUT'])
@jwt_required()
@limiter.limit("10 per minute")
def update_profile():
    """Profil güncelleme"""
    try:
        user_id = get_jwt_identity()
        data = request.get_json()
        
        try:
            user_obj_id = ObjectId(user_id)
        except:
            return jsonify({'error': 'Invalid user ID'}), 401
        
        user = mongo.db.users.find_one({'_id': user_obj_id})
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        update_data = {}
        
        if 'data_source' in data:
            data_source = data['data_source']
            if data_source not in ['csv', 'clockify']:
                return jsonify({'error': 'Invalid data source'}), 400
            
            update_data['data_source'] = data_source
            
            if data_source == 'clockify' and data.get('clockify_api_key'):
                api_key = data['clockify_api_key'].strip()
                if api_key and '*' not in api_key:
                    encrypted_key = encrypt_api_key(api_key)
                    if encrypted_key:
                        update_data['clockify_api_key'] = encrypted_key
        
        if data.get('clockify_api_key') and '*' not in data.get('clockify_api_key', ''):
            api_key = data['clockify_api_key'].strip()
            encrypted_key = encrypt_api_key(api_key)
            if encrypted_key:
                update_data['clockify_api_key'] = encrypted_key
        
        if user['user_type'] == 'individual':
            full_name = sanitize_input(data.get('full_name', ''), 100)
            if full_name:
                update_data['individual_profile.full_name'] = full_name
            if 'phone' in data:
                update_data['individual_profile.phone'] = sanitize_input(data.get('phone', ''), 20)
        
        else:
            company_name = sanitize_input(data.get('company_name', ''), 100)
            if company_name:
                update_data['company_profile.company_name'] = company_name
            if 'contact_person' in data:
                update_data['company_profile.contact_person'] = sanitize_input(data.get('contact_person', ''), 100)
            if 'phone' in data:
                update_data['company_profile.phone'] = sanitize_input(data.get('phone', ''), 20)
            if 'address' in data:
                update_data['company_profile.address'] = sanitize_input(data.get('address', ''), 500)
            
            if 'logo_base64' in data and data['logo_base64']:
                try:
                    logo_data = base64.b64decode(data['logo_base64'].split(',')[1])
                    if len(logo_data) > 2 * 1024 * 1024:
                        return jsonify({'error': 'Logo size must be less than 2MB'}), 400
                    
                    update_data['company_profile.logo_data'] = logo_data
                    update_data['company_profile.logo_mimetype'] = data.get('logo_mimetype', 'image/png')
                except:
                    return jsonify({'error': 'Invalid logo data'}), 400
        
        if update_data:
            mongo.db.users.update_one(
                {'_id': user_obj_id},
                {'$set': update_data}
            )
        
        return jsonify({'message': 'Profile updated successfully'}), 200
        
    except Exception as e:
        return safe_error_response(e, 500)

# ============== CSV PROCESSING ENDPOINTS ==============

@app.route('/api/csv/preview', methods=['POST'])
@jwt_required()
def preview_csv():
    """CSV önizleme"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        df = pd.read_csv(file)
        
        columns = df.columns.tolist()
        sample_data = df.head(10).to_dict('records')
        
        unique_values = {}
        for col in ['Project', 'Client', 'User']:
            if col in df.columns:
                unique_values[col] = df[col].dropna().unique().tolist()
        
        return jsonify({
            'columns': columns,
            'sample_data': sample_data,
            'unique_values': unique_values,
            'total_rows': len(df)
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/csv/convert', methods=['POST'])
@jwt_required()
def convert_csv():
    """CSV'yi Excel'e dönüştür"""
    try:
        user_id = get_jwt_identity()
        user = mongo.db.users.find_one({'_id': ObjectId(user_id)})
        
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        df = pd.read_csv(file)
        
        required_columns = ["Project", "Client", "User", "Start Date", "Duration (h)"]
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return jsonify({'error': f'Missing columns: {", ".join(missing_columns)}'}), 400
        
        if "Billable" not in df.columns:
            df["Billable"] = "No"
        if "Description" not in df.columns:
            df["Description"] = ""
        
        df["Duration (h)"].fillna("00:00:00", inplace=True)
        df["Billable"].fillna("No", inplace=True)
        
        selected_projects = request.form.getlist('projects[]')
        selected_clients = request.form.getlist('clients[]')
        selected_users = request.form.getlist('users[]')
        format_choice = request.form.get('format', 'decimal')
        
        # Kullanıcının seçtiği tarih aralığını al (opsiyonel)
        date_range_start = request.form.get('date_range_start', None)
        date_range_end = request.form.get('date_range_end', None)
        
        if selected_projects and 'all' not in [p.lower() for p in selected_projects]:
            df = df[df["Project"].isin(selected_projects)]
        if selected_clients and 'all' not in [c.lower() for c in selected_clients]:
            df = df[df["Client"].isin(selected_clients)]
        if selected_users and 'all' not in [u.lower() for u in selected_users]:
            df = df[df["User"].isin(selected_users)]
        
        if df.empty:
            return jsonify({'error': 'No data matches filters'}), 400
        
        overall_projects = ", ".join(df["Project"].dropna().unique())
        overall_customers = ", ".join(df["Client"].dropna().unique())
        
        logo_data = None
        company_info = None
        
        if user['user_type'] == 'company':
            profile = user.get('company_profile', {})
            if 'logo_data' in profile:
                logo_data = {
                    'data': profile['logo_data'],
                    'mimetype': profile.get('logo_mimetype', 'image/png')
                }
            company_info = {
                'company_name': profile.get('company_name', ''),
                'contact_person': profile.get('contact_person', ''),
                'phone': profile.get('phone', ''),
                'address': profile.get('address', '')
            }
        
        df['ParsedDate'] = pd.to_datetime(df['Start Date'], format='%d/%m/%Y', errors='coerce')
        
        if not df['ParsedDate'].dropna().empty:
            min_date = df['ParsedDate'].min()
            max_date = df['ParsedDate'].max()
            if min_date.month == max_date.month and min_date.year == max_date.year:
                report_period = min_date.strftime("%B %Y")
            else:
                report_period = f"{min_date.strftime('%B %Y')} - {max_date.strftime('%B %Y')}"
        else:
            report_period = "All Data"
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output = generate_excel_report(
            df, format_choice, report_period,
            overall_projects, overall_customers,
            logo_data, company_info,
            date_range_start=date_range_start,
            date_range_end=date_range_end
        )
        filename = f"Report_{timestamp}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        app.logger.error(f"Error in convert: {str(e)}\n{traceback.format_exc()}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/clockify/workspaces', methods=['GET'])
@jwt_required()
def get_clockify_workspaces():
    """Clockify workspace'lerini getir"""
    try:
        user_id = get_jwt_identity()
        user = mongo.db.users.find_one({'_id': ObjectId(user_id)})
        
        # API key'i header'dan veya database'den al
        api_key = request.headers.get('X-Clockify-Api-Key')
        
        if not api_key or api_key == '':
            # Database'den al ve decrypt et
            encrypted_key = user.get('clockify_api_key', '')
            api_key = decrypt_api_key(encrypted_key) if encrypted_key else None
        
        if not api_key:
            return jsonify({'error': 'Clockify API key required'}), 400
        
        headers = {'X-Api-Key': api_key}
        response = requests.get('https://api.clockify.me/api/v1/workspaces', headers=headers, timeout=10)
        
        if response.status_code != 200:
            return jsonify({'error': 'Invalid Clockify API key'}), 401
        
        workspaces = response.json()
        return jsonify(workspaces), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/clockify/projects', methods=['GET'])
@jwt_required()
def get_clockify_projects():
    """Clockify projelerini getir"""
    try:
        user_id = get_jwt_identity()
        user = mongo.db.users.find_one({'_id': ObjectId(user_id)})
        
        workspace_id = request.args.get('workspace_id')
        
        # API key'i header'dan veya database'den al
        api_key = request.headers.get('X-Clockify-Api-Key')
        
        if not api_key or api_key == '':
            encrypted_key = user.get('clockify_api_key', '')
            api_key = decrypt_api_key(encrypted_key) if encrypted_key else None
        
        if not api_key or not workspace_id:
            return jsonify({'error': 'API key and workspace_id required'}), 400
        
        headers = {'X-Api-Key': api_key}
        response = requests.get(
            f'https://api.clockify.me/api/v1/workspaces/{workspace_id}/projects',
            headers=headers,
            timeout=10
        )
        
        if response.status_code != 200:
            return jsonify({'error': 'Failed to fetch projects'}), 400
        
        return jsonify(response.json()), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/clockify/time-entries', methods=['POST'])
@jwt_required()
def get_clockify_time_entries():
    """Clockify time entries'leri getir ve Excel'e dönüştür"""
    try:
        user_id = get_jwt_identity()
        user = mongo.db.users.find_one({'_id': ObjectId(user_id)})
 
        data = request.get_json()
        workspace_id = data.get('workspace_id')
        start_date   = data.get('start_date')
        end_date     = data.get('end_date')
        project_ids  = data.get('project_ids', [])
 
        api_key = data.get('api_key')
        if not api_key:
            encrypted_key = user.get('clockify_api_key', '')
            api_key = decrypt_api_key(encrypted_key) if encrypted_key else None
 
        if not all([api_key, workspace_id, start_date, end_date]):
            return jsonify({'error': 'Missing required parameters'}), 400
 
        headers = {
            'X-Api-Key': api_key,
            'Content-Type': 'application/json'
        }
 
        try:
            user_response = requests.get(
                'https://api.clockify.me/api/v1/user',
                headers=headers,
                timeout=10
            )
            if user_response.status_code != 200:
                return jsonify({'error': 'Invalid Clockify API key'}), 400
        except Exception as e:
            app.logger.error(f"Auth error: {str(e)}")
            return jsonify({'error': 'Failed to authenticate with Clockify'}), 400
 
        try:
            report_url = f'https://reports.api.clockify.me/v1/workspaces/{workspace_id}/reports/detailed'
            report_payload = {
                "dateRangeStart": start_date,
                "dateRangeEnd":   end_date,
                "detailedFilter": {"page": 1, "pageSize": 1000}
            }
            if project_ids and len(project_ids) > 0:
                report_payload["detailedFilter"]["projects"] = {
                    "ids": project_ids,
                    "contains": "CONTAINS"
                }
 
            report_response = requests.post(report_url, headers=headers, json=report_payload, timeout=30)
            if report_response.status_code != 200:
                app.logger.error(f"API Error: {report_response.text}")
                return jsonify({'error': f'Clockify API error: {report_response.text}'}), 400
 
            report_data  = report_response.json()
            time_entries = report_data.get('timeentries', [])
 
            if not time_entries:
                return jsonify({'error': 'No time entries found'}), 400
 
        except Exception as e:
            app.logger.error(f"Fetch error: {str(e)}\n{traceback.format_exc()}")
            return jsonify({'error': f'Failed to fetch time entries: {str(e)}'}), 400
 
        csv_data = []
        for entry in time_entries:
            try:
                time_interval    = entry.get('timeInterval', {})
                start_str        = time_interval.get('start')
                end_str          = time_interval.get('end')
                duration_seconds = time_interval.get('duration', 0)
 
                if not start_str:
                    continue
 
                start_time = datetime.fromisoformat(start_str.replace('Z', '+00:00'))
                if end_str:
                    end_time = datetime.fromisoformat(end_str.replace('Z', '+00:00'))
                else:
                    end_time = start_time + timedelta(seconds=duration_seconds)
 
                total_seconds = duration_seconds if duration_seconds > 0 else (end_time - start_time).total_seconds()
                hours   = int(total_seconds // 3600)
                minutes = int((total_seconds % 3600) // 60)
                seconds = int(total_seconds % 60)
 
                csv_data.append({
                    'Project':      _safe_str(entry.get('projectName'), 'No Project'),
                    'Client':       _safe_str(entry.get('clientName'),  'No Client'),
                    'User':         _safe_str(entry.get('userName'),    'Unknown'),
                    'Description':  entry.get('description', '') or '',
                    'Start Date':   start_time.strftime('%d/%m/%Y'),
                    'Start Time':   start_time.strftime('%H:%M:%S'),
                    'End Time':     end_time.strftime('%H:%M:%S'),
                    'Duration (h)': f"{hours:02d}:{minutes:02d}:{seconds:02d}",
                    'Billable':     'Yes' if entry.get('billable', False) else 'No'
                })
            except Exception as e:
                app.logger.warning(f"Error processing entry: {str(e)}")
                continue
 
        if not csv_data:
            return jsonify({'error': 'No valid time entries found'}), 400
 
        df = pd.DataFrame(csv_data)
 
        # DataFrame sanitization
        df["Project"]      = df["Project"].apply(lambda x: _safe_str(x, "No Project"))
        df["Client"]       = df["Client"].apply(lambda x: _safe_str(x, "No Client"))
        df["Description"]  = df["Description"].fillna("").astype(str)
        df["Billable"]     = df["Billable"].fillna("No").astype(str)
        df["Duration (h)"] = df["Duration (h)"].fillna("00:00:00").astype(str)
 
        # ── DÜZELTİLMİŞ: sanitization SONRASI hesapla,
        #    "No Project" / "No Client" placeholder'larını filtrele ──
        real_projects  = [p for p in df["Project"].dropna().unique() if p != "No Project"]
        real_customers = [c for c in df["Client"].dropna().unique()  if c != "No Client"]
 
        overall_projects  = ", ".join(real_projects)  if real_projects  else "No Project"
        overall_customers = ", ".join(real_customers) if real_customers else "No Client"
 
        logo_data    = None
        company_info = None
 
        if user and user.get('user_type') == 'company':
            profile = user.get('company_profile', {})
            if 'logo_data' in profile:
                logo_data = {'data': profile['logo_data'], 'mimetype': profile.get('logo_mimetype', 'image/png')}
            company_info = {
                'company_name':   profile.get('company_name', ''),
                'contact_person': profile.get('contact_person', ''),
                'phone':          profile.get('phone', ''),
                'address':        profile.get('address', '')
            }
 
        df['ParsedDate'] = pd.to_datetime(df['Start Date'], format='%d/%m/%Y', errors='coerce')
 
        try:
            req_start = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
            req_end   = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
            if req_start.month == req_end.month and req_start.year == req_end.year:
                report_period = req_start.strftime("%B %Y")
            else:
                report_period = f"{req_start.strftime('%B %Y')} - {req_end.strftime('%B %Y')}"
        except:
            if not df['ParsedDate'].dropna().empty:
                min_date = df['ParsedDate'].min()
                max_date = df['ParsedDate'].max()
                report_period = (
                    min_date.strftime("%B %Y")
                    if (min_date.month == max_date.month and min_date.year == max_date.year)
                    else f"{min_date.strftime('%B %Y')} - {max_date.strftime('%B %Y')}"
                )
            else:
                report_period = "All Data"
 
        format_choice = data.get('format', 'decimal')
        timestamp     = datetime.now().strftime("%Y%m%d_%H%M%S")
 
        output = generate_excel_report(
            df, format_choice, report_period,
            overall_projects, overall_customers,
            logo_data, company_info,
            date_range_start=start_date,
            date_range_end=end_date
        )
        filename = f"Clockify_Report_{timestamp}.xlsx"
 
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
 
    except Exception as e:
        app.logger.error(f"Error: {str(e)}\n{traceback.format_exc()}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/clockify/save-api-key', methods=['POST'])
@jwt_required()
def save_clockify_api_key():
    """Clockify API key'i kaydet - Şifreli"""
    try:
        user_id = get_jwt_identity()
        data = request.get_json()
        
        encrypted_key = encrypt_api_key(data['api_key'])
        
        mongo.db.users.update_one(
            {'_id': ObjectId(user_id)},
            {'$set': {'clockify_api_key': encrypted_key}}
        )
        
        return jsonify({'message': 'API key saved successfully'}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/clockify/get-api-key', methods=['GET'])
@jwt_required()
def get_clockify_api_key():
    """Kayıtlı Clockify API key'i getir - Maskelenmiş"""
    try:
        user_id = get_jwt_identity()
        user = mongo.db.users.find_one({'_id': ObjectId(user_id)})
        
        encrypted_key = user.get('clockify_api_key', '')
        
        if encrypted_key:
            # API key var ama güvenlik için maskelenmiş göster
            decrypted_key = decrypt_api_key(encrypted_key)
            if decrypted_key and len(decrypted_key) > 8:
                # İlk 4 ve son 4 karakteri göster, ortası yıldız
                masked_key = decrypted_key[:4] + '*' * (len(decrypted_key) - 8) + decrypted_key[-4:]
                return jsonify({
                    'api_key': masked_key,
                    'has_key': True
                }), 200
            
        return jsonify({
            'api_key': '',
            'has_key': False
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============== INVOICE ==============
@app.route('/api/invoice/generate', methods=['POST'])
@jwt_required()
def generate_invoice():
    """Invoice oluştur ve Excel'e dönüştür"""
    try:
        user_id = get_jwt_identity()
        user = mongo.db.users.find_one({'_id': ObjectId(user_id)})
        
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        data = request.get_json()
        
        # Gerekli alanları kontrol et
        required_fields = ['invoice_no', 'invoice_date', 'buyer_name', 'services']
        for field in required_fields:
            if field not in data:
                return jsonify({'error': f'Missing field: {field}'}), 400
        
        # Logo ve şirket bilgilerini al
        logo_data = None
        company_info = None
        
        if user.get('user_type') == 'company':
            profile = user.get('company_profile', {})
            if 'logo_data' in profile:
                logo_data = profile['logo_data']
            company_info = {
                'company_name': profile.get('company_name', 'ULEPUS'),
                'address': profile.get('address', 'ODTÜ Teknokent Mustafa Kemal Mah. Dumlupınar Blv. No:280/G İç Kapı No:305 Çankaya/Ankara'),
                'phone': profile.get('phone', '+90-312-486-1158'),
                'email': profile.get('contact_person', 'info@ulepus.com')
            }
        
        # Excel oluştur
        output = generate_invoice_excel(data, logo_data, company_info)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Invoice_{data.get('invoice_no', timestamp)}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        app.logger.error(f"Error in generate_invoice: {str(e)}\n{traceback.format_exc()}")
        return jsonify({'error': str(e)}), 500

# ============== HEALTH CHECK ==============

@app.route('/api/health', methods=['GET'])
@limiter.exempt
def health_check():
    """API sağlık kontrolü"""
    try:
        mongo.db.command('ping')
        return jsonify({
            'status': 'healthy',
            'database': 'connected',
            'timestamp': datetime.utcnow().isoformat()
        }), 200
    except Exception as e:
        return jsonify({
            'status': 'unhealthy',
            'error': 'Database connection failed'
        }), 500

@app.route('/')
def index():
    """Root endpoint"""
    return jsonify({
        'message': 'TimeTracker API',
        'version': '2.0.0',
        'security': 'enhanced',
        'endpoints': {
            'auth': '/api/auth/*',
            'profile': '/api/profile',
            'csv': '/api/csv/*',
            'clockify': '/api/clockify/*',
            'invoice': '/api/invoice/*',
            'health': '/api/health'
        }
    })

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug_mode = os.environ.get('FLASK_ENV') == 'development'
    app.run(debug=debug_mode, host='0.0.0.0', port=port)